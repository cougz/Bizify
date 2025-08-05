import json
import csv
import io
import zipfile
from datetime import datetime
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app import models, schemas, crud


class ExportService:
    def __init__(self, db: Session, user_id: int):
        self.db = db
        self.user_id = user_id
    
    def export_data(self, request: schemas.ExportRequest) -> tuple[bytes, str]:
        """Main export method that delegates to format-specific handlers"""
        if request.format == schemas.ExportFormat.JSON:
            data, filename = self._export_json(request)
            return data, filename
        elif request.format == schemas.ExportFormat.CSV:
            data, filename = self._export_csv(request)
            return data, filename
        elif request.format == schemas.ExportFormat.EXCEL:
            data, filename = self._export_excel(request)
            return data, filename
        elif request.format == schemas.ExportFormat.BACKUP:
            data, filename = self._export_backup(request)
            return data, filename
        else:
            raise ValueError(f"Unsupported export format: {request.format}")
    
    def _get_customers(self, request: schemas.ExportRequest) -> List[models.Customer]:
        """Get customers based on request filters"""
        query = self.db.query(models.Customer).filter(
            models.Customer.user_id == self.user_id
        )
        
        if request.customer_ids:
            query = query.filter(models.Customer.id.in_(request.customer_ids))
        
        return query.all()
    
    def _get_invoices(self, request: schemas.ExportRequest) -> List[models.Invoice]:
        """Get invoices based on request filters"""
        query = self.db.query(models.Invoice).options(
            joinedload(models.Invoice.items),
            joinedload(models.Invoice.customer)
        ).filter(
            models.Invoice.user_id == self.user_id
        )
        
        if request.date_from:
            query = query.filter(models.Invoice.issue_date >= request.date_from)
        
        if request.date_to:
            query = query.filter(models.Invoice.issue_date <= request.date_to)
        
        if request.customer_ids:
            query = query.filter(models.Invoice.customer_id.in_(request.customer_ids))
        
        return query.all()
    
    def _export_json(self, request: schemas.ExportRequest) -> tuple[bytes, str]:
        """Export data as JSON with proper structure"""
        user = crud.get_user(self.db, self.user_id)
        
        data = {
            "export_version": "1.0",
            "export_date": datetime.utcnow().isoformat(),
            "application": "Bizify",
            "user_info": {
                "name": user.name,
                "email": user.email
            }
        }
        
        if request.include_settings:
            settings = crud.get_settings(self.db, self.user_id)
            if settings:
                data["settings"] = {
                    "company_name": settings.company_name,
                    "company_address": settings.company_address,
                    "company_city": settings.company_city,
                    "company_state": settings.company_state,
                    "company_zip": settings.company_zip,
                    "company_country": settings.company_country,
                    "company_phone": settings.company_phone,
                    "company_email": settings.company_email,
                    "company_website": settings.company_website,
                    "tax_rate": settings.tax_rate,
                    "currency": settings.currency,
                    "invoice_prefix": settings.invoice_prefix,
                    "invoice_footer": settings.invoice_footer,
                    "bank_name": settings.bank_name,
                    "bank_iban": settings.bank_iban,
                    "bank_bic": settings.bank_bic,
                    "language": settings.language
                }
        
        if request.include_customers:
            customers = self._get_customers(request)
            data["customers"] = [
                {
                    "name": c.name,
                    "email": c.email,
                    "phone": c.phone,
                    "address": c.address,
                    "city": c.city,
                    "state": c.state,
                    "zip_code": c.zip_code,
                    "country": c.country,
                    "company": c.company,
                    "notes": c.notes
                } for c in customers
            ]
        
        if request.include_invoices:
            invoices = self._get_invoices(request)
            data["invoices"] = [
                {
                    "invoice_number": inv.invoice_number,
                    "customer_email": inv.customer.email,  # Use email as reference
                    "issue_date": inv.issue_date.isoformat() if inv.issue_date else None,
                    "due_date": inv.due_date.isoformat() if inv.due_date else None,
                    "status": inv.status.value,
                    "notes": inv.notes,
                    "tax_rate": inv.tax_rate,
                    "discount": inv.discount,
                    "subtotal": inv.subtotal,
                    "tax_amount": inv.tax_amount,
                    "total": inv.total,
                    "items": [
                        {
                            "description": item.description,
                            "quantity": item.quantity,
                            "unit_price": item.unit_price,
                            "amount": item.amount
                        } for item in inv.items
                    ]
                } for inv in invoices
            ]
        
        json_data = json.dumps(data, indent=2, default=str)
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"bizify_export_{timestamp}.json"
        
        return json_data.encode('utf-8'), filename
    
    def _export_csv(self, request: schemas.ExportRequest) -> tuple[bytes, str]:
        """Export to CSV format with multiple files in a ZIP"""
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            if request.include_customers:
                customers = self._get_customers(request)
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                
                # Headers
                writer.writerow([
                    "Name", "Email", "Company", "Phone", "Address", 
                    "City", "State", "ZIP", "Country", "Notes"
                ])
                
                # Data
                for c in customers:
                    writer.writerow([
                        c.name, c.email, c.company, c.phone, c.address,
                        c.city, c.state, c.zip_code, c.country, c.notes
                    ])
                
                zip_file.writestr("customers.csv", csv_buffer.getvalue())
            
            if request.include_invoices:
                invoices = self._get_invoices(request)
                
                # Invoice summary CSV
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                
                writer.writerow([
                    "Invoice Number", "Customer Name", "Customer Email", 
                    "Issue Date", "Due Date", "Status", "Subtotal", 
                    "Tax Rate", "Tax Amount", "Discount", "Total"
                ])
                
                for inv in invoices:
                    writer.writerow([
                        inv.invoice_number,
                        inv.customer.name,
                        inv.customer.email,
                        inv.issue_date.strftime('%Y-%m-%d') if inv.issue_date else '',
                        inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '',
                        inv.status.value,
                        inv.subtotal,
                        inv.tax_rate,
                        inv.tax_amount,
                        inv.discount,
                        inv.total
                    ])
                
                zip_file.writestr("invoices.csv", csv_buffer.getvalue())
                
                # Invoice items CSV
                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)
                
                writer.writerow([
                    "Invoice Number", "Description", "Quantity", "Unit Price", "Amount"
                ])
                
                for inv in invoices:
                    for item in inv.items:
                        writer.writerow([
                            inv.invoice_number,
                            item.description,
                            item.quantity,
                            item.unit_price,
                            item.amount
                        ])
                
                zip_file.writestr("invoice_items.csv", csv_buffer.getvalue())
        
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"bizify_export_{timestamp}.zip"
        
        return zip_buffer.getvalue(), filename
    
    def _export_excel(self, request: schemas.ExportRequest) -> tuple[bytes, str]:
        """Export to Excel with multiple sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Style definitions
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add sheets based on request
        if request.include_customers:
            ws = wb.create_sheet("Customers")
            headers = ["Name", "Email", "Company", "Phone", "Address", "City", "State", "ZIP", "Country", "Notes"]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            customers = self._get_customers(request)
            for c in customers:
                ws.append([
                    c.name, c.email, c.company, c.phone, c.address,
                    c.city, c.state, c.zip_code, c.country, c.notes
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        if request.include_invoices:
            # Invoice summary sheet
            ws = wb.create_sheet("Invoices")
            headers = [
                "Invoice #", "Customer", "Customer Email", "Issue Date", "Due Date", 
                "Status", "Subtotal", "Tax Rate", "Tax Amount", "Discount", "Total"
            ]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            invoices = self._get_invoices(request)
            for inv in invoices:
                ws.append([
                    inv.invoice_number,
                    inv.customer.name,
                    inv.customer.email,
                    inv.issue_date.strftime('%Y-%m-%d') if inv.issue_date else '',
                    inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '',
                    inv.status.value,
                    inv.subtotal,
                    f"{inv.tax_rate}%",
                    inv.tax_amount,
                    inv.discount,
                    inv.total
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Invoice items sheet
            ws = wb.create_sheet("Invoice Items")
            headers = ["Invoice #", "Description", "Quantity", "Unit Price", "Amount"]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            for inv in invoices:
                for item in inv.items:
                    ws.append([
                        inv.invoice_number,
                        item.description,
                        item.quantity,
                        item.unit_price,
                        item.amount
                    ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        if request.include_settings:
            settings = crud.get_settings(self.db, self.user_id)
            if settings:
                ws = wb.create_sheet("Company Settings")
                ws.append(["Setting", "Value"])
                
                # Style headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                settings_data = [
                    ("Company Name", settings.company_name),
                    ("Address", settings.company_address),
                    ("City", settings.company_city),
                    ("State", settings.company_state),
                    ("ZIP Code", settings.company_zip),
                    ("Country", settings.company_country),
                    ("Phone", settings.company_phone),
                    ("Email", settings.company_email),
                    ("Website", settings.company_website),
                    ("Tax Rate", f"{settings.tax_rate}%"),
                    ("Currency", settings.currency),
                    ("Invoice Prefix", settings.invoice_prefix),
                    ("Bank Name", settings.bank_name),
                    ("Bank IBAN", settings.bank_iban),
                    ("Bank BIC", settings.bank_bic),
                    ("Language", settings.language),
                ]
                
                for setting, value in settings_data:
                    ws.append([setting, value])
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 40
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"bizify_export_{timestamp}.xlsx"
        
        return buffer.getvalue(), filename
    
    def _export_backup(self, request: schemas.ExportRequest) -> tuple[bytes, str]:
        """Export complete backup with all data and metadata"""
        # For backup, always include everything
        backup_request = schemas.ExportRequest(
            format=schemas.ExportFormat.JSON,
            include_customers=True,
            include_invoices=True,
            include_settings=True
        )
        
        json_data, _ = self._export_json(backup_request)
        
        # Create ZIP with additional metadata
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Add main data file
            zip_file.writestr("bizify_data.json", json_data)
            
            # Add metadata
            metadata = {
                "backup_date": datetime.utcnow().isoformat(),
                "version": "1.0",
                "application": "Bizify",
                "backup_type": "complete",
                "user_id": self.user_id,
                "total_customers": len(self._get_customers(backup_request)),
                "total_invoices": len(self._get_invoices(backup_request))
            }
            
            zip_file.writestr("metadata.json", json.dumps(metadata, indent=2))
            
            # Add README
            readme_content = """# Bizify Backup

This is a complete backup of your Bizify data.

## Contents:
- bizify_data.json: Complete data export
- metadata.json: Backup information
- README.md: This file

## Restore Instructions:
1. Access Bizify settings
2. Go to Import/Export section
3. Select this backup file
4. Choose import options
5. Click Import

Generated on: """ + datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
            
            zip_file.writestr("README.md", readme_content)
        
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"bizify_backup_{timestamp}.zip"
        
        return zip_buffer.getvalue(), filename

    def get_media_type(self, format: schemas.ExportFormat) -> str:
        """Get the appropriate media type for the export format"""
        media_types = {
            schemas.ExportFormat.JSON: "application/json",
            schemas.ExportFormat.CSV: "application/zip",
            schemas.ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            schemas.ExportFormat.BACKUP: "application/zip"
        }
        return media_types.get(format, "application/octet-stream")